
from openpyxl import Workbook
from datetime import datetime

wb = Workbook()

ws = wb.active

ws.title = "Ajantha"
# cells
ws['B9'] = "Greetings Mr.Sachin"

ws['D9'] = 1048576

ws['F9'] = datetime.now()

ws['H6'] = 75
ws['H7'] = 100
ws['H8'] = 250
ws['H9'] = "=sum(H6:H8)"

wb.save("FirstExcel.xlsx")