
from openpyxl import load_workbook

wb = load_workbook("FirstExcel.xlsx")

wb.active = wb['Players']

ws = wb.active

for row in ws.iter_rows(min_row=7, max_row=11, min_col=1, max_col=5):
    for cell in row:
        # if cell.value == "Ponting":
        #     cell.value = "ricky ponting".upper()
        if cell.column == 1 or cell.column == 5:
            print(cell.value, end=" ")
    print()

wb.save("FirstExcel.xlsx")
