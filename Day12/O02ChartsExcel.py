
from openpyxl import  load_workbook
from openpyxl.chart import PieChart3D, Reference
from openpyxl.chart.label import DataLabelList

wb = load_workbook("FirstExcel.xlsx")

wb.create_sheet("Chart")

wb.active = wb["Chart"]

ws = wb.active

data = [
    ('Types of Expenses', "Amount Spent"),
    ("Grocery", 300),
    ("Electricity", 150),
    ("Child tution", 125),
    ("House keeping", 35),
    ("gardening", 70),
    ("Misl Expense", 500)
]

for row in data:
    ws.append(row)

pie = PieChart3D()

labels = Reference(ws, min_col=1, min_row=2, max_row=7)
data = Reference(ws, min_col=2, min_row = 1, max_row=7)
pie.add_data(data, titles_from_data = True)
pie.set_categories(labels)
pie.title = "Expenditure Pie chart"
pie.dataLabels = DataLabelList()
pie.dataLabels.showVal = True

ws.add_chart(pie, "G10")
wb.save("FirstExcel.xlsx")
